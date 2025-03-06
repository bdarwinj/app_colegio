# src/utils/export_students.py
import os
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from itertools import groupby

# Clase para PDF con numeración en el pie de página
class PDFWithFooter(FPDF):
    def footer(self):
        self.set_y(-15)  # 15 mm desde el final de la página
        self.set_font("Arial", "I", 8)
        self.cell(0, 10, f"Página {self.page_no()}", 0, 0, "R")

# Constantes para los encabezados de la tabla de estudiantes
HEADERS = ["Identificacion", "Nombre", "Apellido", "Grado", "Representante", "Numero de Telefono"]

def _preparar_students(students):
    """
    Convierte cada registro de estudiante a diccionario y los ordena por curso,
    siguiendo la secuencia definida (de pre-jardin a once). Si existen cursos con
    el mismo nombre base, se ordenan alfabéticamente por el nombre completo.
    """
    prepared_students = []
    keys = ["id", "identificacion", "nombre", "apellido", "course_id", "representante", "telefono", "active"]
    
    for student in students:
        if isinstance(student, dict):
            prepared_students.append(student.copy())
        else:
            try:
                student_dict = dict(zip(keys, student))
                prepared_students.append(student_dict)
            except Exception as e:
                print(f"Error al convertir estudiante a diccionario: {e}")
                continue

    # Secuencia de cursos (en minúsculas)
    course_order = ["pre-jardin", "jardin", "transicion", "primero", "segundo", "tercero",
                    "cuarto", "quinto", "sexto", "septimo", "octavo", "noveno", "decimo", "once"]
    
    def course_sort_key(student):
        course = student.get("course_name", "")
        base = course.split("-")[0].strip().lower() if course else ""
        try:
            rank = course_order.index(base)
        except ValueError:
            rank = len(course_order)
        return (rank, course.lower())
    
    return sorted(prepared_students, key=course_sort_key)

def export_students_to_excel(students, output_filename, school_name, logo_path, course_controller):
    """
    Exporta estudiantes a un archivo Excel con hojas separadas por curso y una hoja 'Todos',
    con un diseño visualmente atractivo.
    """
    wb = openpyxl.Workbook()
    students_sorted = _preparar_students(students)
    generation_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    def get_course_name(student):
        """
        Obtiene el nombre del curso para un estudiante usando 'course_name' si existe,
        o consultando course_controller con 'course_id'.
        """
        if "course_name" in student and student["course_name"]:
            return student["course_name"]
        course_id = student.get("course_id", "")
        if course_id:
            try:
                course_data = course_controller.get_course_by_id(course_id)
                if course_data:
                    name = course_data.get("name", "")
                    seccion = course_data.get("seccion", "")
                    return f"{name} - {seccion.strip()}" if seccion and seccion.strip() else name
                return ""
            except Exception as e:
                print(f"Error al obtener nombre del curso para course_id {course_id}: {e}")
        return ""
    
    def _crear_hoja(wb, sheet_name, data):
        ws = wb.create_sheet(title=sheet_name)
        row_offset = 1

        # Encabezado con logo y título
        if os.path.exists(logo_path):
            try:
                img = XLImage(logo_path)
                img.height = 80
                img.width = 80
                ws.add_image(img, "A1")
            except Exception as e:
                print(f"Error al insertar logo en Excel: {e}")
        
        ws.merge_cells(start_row=row_offset, start_column=2, end_row=row_offset, end_column=6)
        header_cell = ws.cell(row=row_offset, column=2, value=school_name)
        header_cell.font = Font(bold=True, size=16, color="003366")  # Azul oscuro
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid")  # Azul claro
        row_offset += 1

        # Subtítulo y fecha
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=6)
        subtitle_cell = ws.cell(row=row_offset, column=1, value=f"Reporte de Estudiantes - Generado el {generation_date}")
        subtitle_cell.font = Font(size=10, italic=True, color="646464")  # Gris
        subtitle_cell.alignment = Alignment(horizontal="center", vertical="center")
        row_offset += 2

        # Encabezados de la tabla
        ws.append(HEADERS)
        header_row = ws.max_row
        header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Azul oscuro
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                             top=Side(style="thin"), bottom=Side(style="thin"))
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")  # Blanco
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Filas de datos
        fill = False
        for st in data:
            course_name_val = get_course_name(st)
            row_data = [
                st.get("identificacion", ""),
                st.get("nombre", ""),
                st.get("apellido", ""),
                course_name_val,
                st.get("representante", ""),
                st.get("telefono", "")
            ]
            ws.append(row_data)
            fill = not fill
            for col in range(1, len(HEADERS) + 1):
                cell = ws.cell(row=ws.max_row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid") if fill else PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Azul claro / Blanco
        
        # Ajustar ancho de columnas dinámicamente
        for col in range(1, len(HEADERS) + 1):
            max_length = max(len(str(cell.value or "")) for cell in ws[get_column_letter(col)])
            ws.column_dimensions[get_column_letter(col)].width = max_length + 5
        
        # Pie de página simulado
        row_offset = ws.max_row + 2
        ws.merge_cells(start_row=row_offset, start_column=1, end_row=row_offset, end_column=6)
        footer_cell = ws.cell(row=row_offset, column=1, value=f"Total de estudiantes: {len(data)} - Reporte generado el {generation_date}")
        footer_cell.font = Font(size=8, italic=True, color="969696")  # Gris claro
        footer_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        return ws
    
    # Agrupar estudiantes por curso
    grouped_by_course = groupby(students_sorted, key=lambda s: get_course_name(s))
    for course_name, group in grouped_by_course:
        group_list = list(group)
        sheet_title = f"{course_name}" if course_name else "Grado_Desconocido"
        _crear_hoja(wb, sheet_title, group_list)
    
    default_sheet = wb.active
    wb.remove(default_sheet)
    _crear_hoja(wb, "Todos", students_sorted)
    
    wb.save(output_filename)
    return output_filename

class PDFWithHeaderFooter(FPDF):
    def __init__(self, school_name, logo_path=None):
        super().__init__(orientation="L", unit="mm", format=(216, 385))  # Hoja horizontal
        self.school_name = school_name
        self.logo_path = logo_path

    def header(self):
        # Logo (opcional, si se proporciona)
        if self.logo_path and os.path.exists(self.logo_path):
            try:
                self.image(self.logo_path, x=10, y=8, w=25)
            except Exception as e:
                print(f"Error al insertar logo: {e}")
        # Nombre del colegio al lado del logo, centrado en la hoja
        self.set_font("Arial", "B", 18)
        self.set_text_color(0, 51, 102)  # Azul oscuro
        self.set_xy(0, 10)
        self.cell(0, 10, self.school_name, ln=True, align="C")
        
        # Fecha de generación
        self.set_font("Arial", "I", 10)
        self.set_text_color(100, 100, 100)  # Gris
        self.cell(0, 5, f"Generado el {datetime.datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", ln=True, align="C")
        
        # Línea decorativa
        self.set_line_width(0.5)
        self.set_draw_color(0, 51, 102)  # Azul oscuro
        self.line(10, 35, self.w - 10, 35)
        self.ln(10)

    def footer(self):
        self.set_y(-15)
        self.set_font("Arial", "I", 8)
        self.set_text_color(150, 150, 150)  # Gris claro
        self.cell(0, 10, f"Página {self.page_no()} - Reporte generado por Sistema Escolar", 0, 0, "C")

def export_students_to_pdf(students, output_filename, school_name, logo_path=None, course_controller=None):
    """
    Exporta estudiantes a un archivo PDF ordenados por curso, con un diseño visualmente atractivo.
    
    :param students: Lista de estudiantes (diccionarios con datos).
    :param output_filename: Nombre del archivo PDF a generar.
    :param school_name: Nombre del colegio para el encabezado.
    :param logo_path: Ruta opcional al logo del colegio.
    :param course_controller: Controlador de cursos (no usado aquí, pero mantenido por compatibilidad).
    :return: Ruta del archivo PDF generado.
    """
    pdf = PDFWithHeaderFooter(school_name, logo_path)
    pdf.set_margins(left=15, top=40, right=15)  # Márgenes amplios
    pdf.add_page()

    # Título de la tabla
    pdf.set_font("Arial", "B", 14)
    pdf.set_text_color(0, 51, 102)  # Azul oscuro
    pdf.cell(0, 10, "Lista de Estudiantes", ln=True, align="C")
    pdf.ln(5)

    # Preparar estudiantes (asumiendo que tienes esta función definida)
    students_sorted = _preparar_students(students)

    # Definir encabezados (ajusta según tu código original)
    HEADERS = ["Identificación", "Nombre", "Apellido", "Curso", "Representante", "Teléfono"]
    
    # Calcular anchos de columna dinámicamente
    col_widths = [max([pdf.get_string_width(str(st.get(key, ""))) for st in students_sorted]) + 8 
                  for key in ["identificacion", "nombre", "apellido", "course_name", "representante", "telefono"]]
    total_width = sum(col_widths)
    if total_width > pdf.w - 30:
        scale_factor = (pdf.w - 30) / total_width
        col_widths = [w * scale_factor for w in col_widths]

    # Encabezados de la tabla
    pdf.set_fill_color(0, 102, 204)  # Azul claro
    pdf.set_text_color(255, 255, 255)  # Blanco
    pdf.set_font("Arial", "B", 10)
    for i, header in enumerate(HEADERS):
        pdf.cell(col_widths[i], 8, header, border=1, align="C", fill=True)
    pdf.ln()

    # Filas de la tabla
    pdf.set_font("Arial", "", 9)
    pdf.set_text_color(0, 0, 0)  # Negro
    fill = False
    for st in students_sorted:
        pdf.set_fill_color(240, 245, 255) if fill else pdf.set_fill_color(255, 255, 255)  # Azul claro / Blanco
        row_data = [
            str(st.get("identificacion", "")),
            str(st.get("nombre", "")),
            str(st.get("apellido", "")),
            st.get("course_name", ""),
            str(st.get("representante", "")),
            str(st.get("telefono", ""))
        ]
        for i, data in enumerate(row_data):
            pdf.cell(col_widths[i], 8, data, border=1, align="C", fill=True)
        pdf.ln()
        fill = not fill

    # Guardar el PDF
    pdf.output(output_filename)
    return output_filename