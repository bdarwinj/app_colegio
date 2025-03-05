# src/utils/import_students.py
import openpyxl
import re
import unicodedata
import datetime
from src.logger import logger
from src.controllers.enrollment_controller import EnrollmentController

def normalize_string(s):
    """
    Normaliza una cadena: quita acentos, la convierte a minúsculas y elimina espacios extra.
    """
    if not s:
        return ""
    normalized = unicodedata.normalize('NFKD', s).encode('ASCII', 'ignore').decode('utf-8')
    return normalized.lower().strip()

def import_students_from_excel(db, excel_path, course_controller, student_controller):
    """
    Importa estudiantes desde un archivo Excel y, para cada estudiante registrado, crea automáticamente una inscripción.
    
    Cada hoja representa un curso y su nombre se usa para identificar el curso.
    Se esperan las siguientes columnas en cada hoja (la primera fila contiene los encabezados):
      - "nombre del estudiante" o "nombre": contiene 4 palabras; las dos primeras serán el nombre y las dos siguientes el apellido.
      - "identificación": número de identificación (si falta, se asigna "123456789")
      - "telefono": número de teléfono
      - "correo electronico" o "correo electrónico": se valida que sea un correo; si falta, se asigna "email@email.com"
      - "acudiente": nombre del representante.
    
    Las comparaciones de encabezados se realizan normalizando (sin mayúsculas, tildes, etc.).
    Además, al buscar el curso correspondiente, se compara el nombre y la sección (permitiendo "PRIMERO A" o "PRIMERO - A").
    
    Además, para cada estudiante registrado se crea una inscripción automática en el año actual con estado "inscrito".
    
    Retorna una tupla (num_importados, errores), donde errores es una lista de mensajes.
    """
    errors = []
    imported_count = 0
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        logger.exception("Error al cargar el archivo Excel")
        return 0, [f"Error al cargar el archivo Excel: {e}"]
    
    # Definir los nombres requeridos (normalizados) y sus alternativas
    required_columns = {
        "nombre": ["nombre del estudiante", "nombre"],
        "identificación": ["identificación"],
        "telefono": ["telefono"],
        "correo electronico": ["correo electronico", "correo electrónico"],
        "acudiente": ["acudiente"]
    }
    
    # Normalizar las alternativas para cada campo
    for key in required_columns:
        required_columns[key] = [normalize_string(alt) for alt in required_columns[key]]
    
    # Procesar cada hoja
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Leer encabezados de la primera fila y normalizarlos
        headers = {}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        for idx, cell in enumerate(header_row):
            if cell:
                headers[normalize_string(cell)] = idx
        
        # Verificar que existan todas las columnas requeridas
        missing_cols = []
        for key, alternatives in required_columns.items():
            if not any(alt in headers for alt in alternatives):
                missing_cols.append(key)
        if missing_cols:
            errors.append(f"Hoja '{sheet_name}': Faltan columnas para: {', '.join(missing_cols)}")
            continue
        
        # Obtener índices de las columnas usando la primera alternativa encontrada para cada campo
        col_indices = {}
        for key, alternatives in required_columns.items():
            for alt in alternatives:
                if alt in headers:
                    col_indices[key] = headers[alt]
                    break
        
        # Buscar el curso en la base de datos usando el nombre de la hoja.
        # Se normaliza el nombre de la hoja y se intenta separar en dos partes (base y sección),
        # excepto si la hoja es "pre-jardin", en cuyo caso se omite la división.
        course_data = None
        courses = course_controller.get_all_courses()
        sheet_name_norm = normalize_string(sheet_name)
        if sheet_name_norm == "pre-jardin":
            base_sheet = "pre-jardin"
            section_sheet = ""
        else:
            if "-" in sheet_name_norm:
                parts = [p.strip() for p in sheet_name_norm.split("-") if p.strip()]
            else:
                parts = sheet_name_norm.split()
            base_sheet = parts[0] if parts else sheet_name_norm
            section_sheet = parts[1] if len(parts) > 1 else ""
        # Buscar en la lista de cursos
        for course in courses:
            course_name_norm = normalize_string(course.get("name", ""))
            course_seccion_norm = normalize_string(course.get("seccion", ""))
            if course_name_norm == base_sheet:
                if section_sheet:
                    if course_seccion_norm == section_sheet:
                        course_data = course
                        break
                else:
                    course_data = course
                    break
        if not course_data:
            errors.append(f"Hoja '{sheet_name}': Curso no encontrado en la base de datos.")
            continue
        course_id = course_data.get("id")
        
        # Crear instancia de EnrollmentController para registrar inscripciones
        from src.controllers.enrollment_controller import EnrollmentController
        enrollment_controller = EnrollmentController(db, student_controller, course_controller)
        current_year = datetime.datetime.now().year
        
        # Procesar filas (desde la segunda fila)
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                nombre_full = row[col_indices["nombre"]]
                if not nombre_full:
                    errors.append(f"Hoja '{sheet_name}': Fila con 'nombre' vacío, omitiendo.")
                    continue
                words = str(nombre_full).split()
                if len(words) < 4:
                    errors.append(f"Hoja '{sheet_name}': Valor en 'nombre' incorrecto: {nombre_full}")
                    continue
                # Las dos primeras palabras son el nombre; las dos siguientes, el apellido
                nombre = " ".join(words[2:4])
                apellido = " ".join(words[:2])
                identificacion = str(row[col_indices["identificación"]]).strip() if row[col_indices["identificación"]] else "123456789"
                telefono = str(row[col_indices["telefono"]]).strip() if row[col_indices["telefono"]] else ""
                email = str(row[col_indices["correo electronico"]]).strip() if row[col_indices["correo electronico"]] else "email@email.com"
                acudiente = str(row[col_indices["acudiente"]]).strip() if row[col_indices["acudiente"]] else ""
                
                # Validar formato de email
                if email and not re.match(r"[^@]+@[^@]+\.[^@]+", email):
                    errors.append(f"Hoja '{sheet_name}': Email inválido: {email}")
                    continue

                success, msg = student_controller.register_student(
                    identificacion, nombre, apellido, course_id, acudiente, telefono, email
                )
                if success:
                    # Obtener el estudiante recién registrado para crear su inscripción
                    student_record = student_controller.get_student_by_identification(identificacion)
                    if student_record:
                        student_id = student_record["id"]
                        enroll_success, enroll_msg, enrollment_id = enrollment_controller.create_enrollment(
                            student_id, course_id, current_year, status="inscrito"
                        )
                        if not enroll_success:
                            errors.append(f"Hoja '{sheet_name}', ID {identificacion}: Error en inscripción: {enroll_msg}")
                    else:
                        errors.append(f"Hoja '{sheet_name}', ID {identificacion}: No se pudo recuperar el estudiante recién creado.")
                    imported_count += 1
                else:
                    errors.append(f"Hoja '{sheet_name}', ID {identificacion}: {msg}")
            except Exception as e:
                logger.exception(f"Error procesando fila en hoja '{sheet_name}'")
                errors.append(f"Hoja '{sheet_name}', fila desconocida: {e}")
    return imported_count, errors
