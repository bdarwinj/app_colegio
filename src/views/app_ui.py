import tkinter as tk
from tkinter import messagebox, filedialog, ttk
import datetime
import logging
import os
import openpyxl
from fpdf import FPDF  # Biblioteca para generar PDFs
from src.controllers.enrollment_controller import EnrollmentController  # Controlador de inscripciones
from src.controllers.student_controller import StudentController
from src.controllers.course_controller import CourseController
from src.controllers.config_controller import ConfigController
from src.controllers.user_controller import UserController
from src.controllers.payment_controller import PaymentController
from src.views.enrollment_management_ui import EnrollmentManagementUI
from src.views.config_ui import ConfigUI
from src.views.user_management_ui import UserManagementUI
from src.views.payment_ui import PaymentUI
from src.views.login_ui import LoginUI
from src.views.student_details_window import StudentDetailsWindow
from src.views.header_frame import HeaderFrame
from src.views.admin_panel import AdminPanel
from src.utils.import_students import import_students_from_excel
from src.views.dashboard_window import DashboardWindow
from src.views.student_registration_frame import StudentRegistrationFrame
from src.views.students_list_frame import StudentsListFrame
from src.views.action_buttons import ActionButtons
from src.views.backup_restore_frame import BackupRestoreFrame
from src.views.change_password_window import ChangePasswordWindow
from src.views.course_management_window import CourseManagementWindow
from src.utils.export_students import export_students_to_excel, export_students_to_pdf
from src.utils.backup_restore import backup_database, restore_database
from src.utils.pdf_utils import PDFWithHeaderFooter

# Configuración del logging
logging.basicConfig(filename='app_ui.log', level=logging.INFO, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Constantes para mensajes
MSG_ERROR = "Error"
MSG_SUCCESS = "Éxito"
MSG_FIELDS_INCOMPLETE = "Campos incompletos"
MSG_INVALID_VALUE = "Valor inválido"
MSG_SELECTION_REQUIRED = "Selección requerida"
MSG_CONFIRMATION = "¿Está seguro de cerrar la sesión?"

class AppUI:
    def __init__(self, db, user):
        self.db = db
        self.user = user
        self.student_controller = StudentController(self.db)
        self.course_controller = CourseController(self.db)
        self.config_controller = ConfigController(self.db)
        self.user_controller = UserController(self.db)
        self.payment_controller = PaymentController(self.db)
        self.root = tk.Tk()

        configs = self.config_controller.get_all_configs()
        self.school_name = configs.get("SCHOOL_NAME") or "School Name"
        self.logo_path = configs.get("LOGO_PATH") or ""
        self.abs_logo_path = os.path.abspath(self.logo_path)

        self.root.title(f"{self.school_name} - Sistema de Pagos (Usuario: {self.user.username})")
        self.root.state("zoomed")
        self.create_widgets()

    def import_students_excel(self):
        # Abre un cuadro de diálogo para seleccionar un archivo Excel
        excel_path = filedialog.askopenfilename(
            title="Seleccionar archivo Excel",
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if not excel_path:
            return
        try:
            imported_count, errors = import_students_from_excel(self.db, excel_path, 
                                                                self.course_controller, 
                                                                self.student_controller)
            message = f"Estudiantes importados: {imported_count}\n"
            if errors:
                message += "Errores:\n" + "\n".join(errors)
            else:
                message += "Importación completada sin errores."
            messagebox.showinfo("Resultado de Importación", message)
            self.refrescar_lista()
        except Exception as e:
            logging.error(f"Error al importar estudiantes desde Excel: {e}")
            messagebox.showerror(MSG_ERROR, f"Error al importar estudiantes: {str(e)}")

    def create_widgets(self):
        # Crea la barra de encabezado
        header_frame = HeaderFrame(self.root, self.school_name, self.abs_logo_path, 
                                   self.open_change_password_window, self.logout)
        header_frame.pack(fill="x", padx=10, pady=10)

        if self.user.role == "admin":
            self.frame_admin = AdminPanel(
                self.root,
                config_command=self.editar_configuracion,
                payment_command=self.registrar_pago,
                courses_command=self.manage_courses,
                users_command=self.manage_users,
                dashboard_command=self.open_dashboard_window,
                import_command=self.import_students_excel
            )
            self.frame_admin.pack(padx=10, pady=10, fill="x")
            self.frame_form = StudentRegistrationFrame(self.root, self.course_controller, 
                                                       self.registrar_estudiante)
            self.frame_form.pack(fill="both", expand=True, padx=5, pady=5)
            self.frame_form.populate_courses()
        elif self.user.role == "user":
            self.btn_registrar_pago = ttk.Button(self.root, text="Registrar Pago", 
                                                 command=self.registrar_pago)
            self.btn_registrar_pago.pack(pady=5)

        self.frame_lista = StudentsListFrame(self.root, self.student_controller, 
                                             self.on_student_double_click)
        self.frame_lista.pack(padx=10, pady=10, fill="both", expand=True)

        actions_frame = ActionButtons(self.root, self.refrescar_lista, self.generar_pdf, 
                                      self.export_students_excel, self.export_students_pdf)
        actions_frame.pack(pady=5)

        if self.user.role == "admin":
            backup_restore_frame = BackupRestoreFrame(self.root, self.backup_database, 
                                                      self.restore_database, self.manage_enrollments)
            backup_restore_frame.pack(pady=5)

    def open_dashboard_window(self):
        # Abre la ventana del tablero
        DashboardWindow(self.root, self.db, self.student_controller, 
                        self.course_controller, self.payment_controller)

    def editar_configuracion(self):
        # Abre la interfaz de configuración
        ConfigUI(self.db)

    def registrar_pago(self):
        # Abre la interfaz de registro de pagos
        PaymentUI(self.db)

    def manage_courses(self):
        # Abre la ventana de gestión de cursos
        CourseManagementWindow(self.root, self.course_controller)

    def manage_users(self):
        # Abre la interfaz de gestión de usuarios
        UserManagementUI(self.db)

    def registrar_estudiante(self):
        # Obtiene los datos del formulario
        identificacion = self.frame_form.entries["Número de Identificación"].get()
        nombre = self.frame_form.entries["Nombre"].get()
        apellido = self.frame_form.entries["Apellido"].get()
        email = self.frame_form.entries["Correo Electrónico"].get()
        representante = self.frame_form.entries["Representante"].get()
        telefono = self.frame_form.entries["Teléfono"].get()
        course_name = self.frame_form.combo_course.get()

        if not (identificacion and nombre and apellido and email and representante and telefono and course_name):
            messagebox.showwarning(MSG_FIELDS_INCOMPLETE, "Por favor, llene todos los campos.")
            return

        course_id = self.frame_form.course_map.get(course_name)
        success, msg = self.student_controller.register_student(
            identificacion, nombre, apellido, course_id, representante, telefono, email
        )
        if success:
            messagebox.showinfo(MSG_SUCCESS, msg)
            student_record = self.student_controller.get_student_by_identification(identificacion)
            if student_record:
                student_id = student_record["id"]
                current_year = datetime.datetime.now().year
                enrollment_controller = EnrollmentController(self.db, self.student_controller, 
                                                             self.course_controller)
                enroll_success, enroll_msg, enrollment_id = enrollment_controller.create_enrollment(
                    student_id, course_id, current_year, status="inscrito"
                )
                if enroll_success:
                    logging.info(f"Inscripción creada correctamente para el estudiante {identificacion} (ID: {enrollment_id}).")
                else:
                    logging.error(f"Error al crear inscripción para el estudiante {identificacion}: {enroll_msg}")
            else:
                logging.error("No se pudo recuperar el registro del estudiante recién creado.")
            self.limpiar_formulario()
            self.refrescar_lista()
        else:
            messagebox.showerror(MSG_ERROR, msg)

    def limpiar_formulario(self):
        # Limpia todos los campos del formulario
        for entry in self.frame_form.entries.values():
            entry.delete(0, tk.END)
        self.frame_form.combo_course.set("")

    def refrescar_lista(self):
        # Actualiza la lista de estudiantes
        self.frame_lista.all_students = self.student_controller.get_all_students()
        self.frame_lista.populate_student_list(self.frame_lista.all_students)

    def on_student_double_click(self, event):
        # Maneja el doble clic en un estudiante de la lista
        try:
            selected = self.frame_lista.tree.selection()
            if selected:
                item = self.frame_lista.tree.item(selected[0])
                student_identificacion = item["values"][1]
                StudentDetailsWindow(self.db, student_identificacion)
        except Exception as e:
            logging.error(f"Error al abrir detalles del estudiante: {e}")
            messagebox.showerror(MSG_ERROR, f"Error al abrir los detalles del estudiante: {str(e)}")

    def generar_pdf(self):
        # Genera un PDF de paz y salvo para el estudiante seleccionado
        selected = self.frame_lista.tree.selection()
        if not selected:
            messagebox.showwarning("Selección requerida", "Seleccione un estudiante para generar el PDF")
            return
        
        item = self.frame_lista.tree.item(selected[0])
        estudiante_data = item["values"]
        
        pdf = PDFWithHeaderFooter(self.abs_logo_path, self.school_name, receipt_number="", titulo="Paz y Salvo")
        pdf.set_margins(left=15, top=40, right=15)
        pdf.add_page()
        
        # Sección de Datos del Estudiante
        pdf.set_font("Arial", "B", 12)
        pdf.set_text_color(0, 51, 102)  # Azul oscuro
        pdf.cell(0, 10, "Datos del Estudiante", ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(0, 0, 0)  # Negro
        
        # Fondo gris claro para los detalles
        pdf.set_fill_color(240, 240, 240)  # Gris claro
        campos = ["ID", "Identificación", "Nombre", "Apellido", "Curso"]
        for idx, campo in enumerate(campos):
            pdf.cell(50, 8, f"{campo}:", border=1, align="L", fill=True)
            pdf.cell(130, 8, str(estudiante_data[idx]), border=1, align="L", ln=True, fill=True)
        pdf.ln(10)
        
        # Sección de Total Pagado y Fecha de Emisión
        pdf.set_font("Arial", "B", 12)
        pdf.set_text_color(0, 51, 102)  # Azul oscuro
        pdf.cell(0, 10, "Estado de Pago", ln=True)
        pdf.set_font("Arial", "", 10)
        pdf.set_text_color(0, 0, 0)  # Negro
        
        student_id = estudiante_data[0]
        pagos = self.payment_controller.get_payments_by_student(student_id)
        total_pagado = sum(float(payment["amount"]) for payment in pagos if payment["amount"] is not None)
        formatted_total = f"${total_pagado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        pdf.set_fill_color(240, 240, 240)  # Gris claro
        pdf.cell(50, 8, "Total Pagado:", border=1, align="L", fill=True)
        pdf.set_text_color(0, 102, 204)  # Azul claro
        pdf.cell(130, 8, formatted_total, border=1, align="L", ln=True, fill=True)
        pdf.set_text_color(0, 0, 0)  # Negro
        pdf.cell(50, 8, "Fecha de Emisión:", border=1, align="L", fill=True)
        pdf.cell(130, 8, datetime.date.today().strftime("%d/%m/%Y"), border=1, align="L", ln=True, fill=True)
        default_filename = f"paz_y_salvo_{estudiante_data[2]}.pdf"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf")],
            initialfile=default_filename,
            title="Guardar Paz y Salvo"
        )
        if file_path:
            pdf.output(file_path)
            messagebox.showinfo("PDF generado", f"El PDF '{file_path}' ha sido generado correctamente.")

    def _export_students(self, export_func, file_extension, file_type):
        # Exporta estudiantes a un archivo (Excel o PDF)
        try:
            estudiantes = self.student_controller.get_all_students()
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            default_filename = f"{self.school_name}_Listado_Estudiantes_{timestamp}{file_extension}"
            file_path = filedialog.asksaveasfilename(defaultextension=file_extension,
                                                     filetypes=[(file_type, f"*{file_extension}")],
                                                     initialfile=default_filename)
            if not file_path:
                return
            export_func(estudiantes, file_path, self.school_name, self.logo_path, self.course_controller)
            messagebox.showinfo("Exportación exitosa", f"Listado exportado a {file_type}: {file_path}")
        except Exception as e:
            logging.error(f"Error al exportar a {file_type}: {e}")
            messagebox.showerror(MSG_ERROR, f"Error al exportar a {file_type}: {str(e)}")

    def export_students_excel(self):
        # Exporta estudiantes a Excel
        self._export_students(export_students_to_excel, ".xlsx", "Excel")

    def export_students_pdf(self):
        # Exporta estudiantes a PDF
        self._export_students(export_students_to_pdf, ".pdf", "PDF")

    def logout(self):
        # Cierra la sesión del usuario
        confirm = messagebox.askyesno("Cerrar Sesión", MSG_CONFIRMATION)
        if confirm:
            self.root.destroy()
            LoginUI(self.db).run()

    def open_change_password_window(self):
        # Abre la ventana para cambiar contraseña
        ChangePasswordWindow(self.root, self.user_controller, self.user.username)

    def manage_enrollments(self):
        # Abre la interfaz de gestión de inscripciones
        EnrollmentManagementUI(self.root, self.student_controller, self.course_controller)

    def backup_database(self):
        # Realiza un respaldo de la base de datos
        try:
            backup_file = backup_database()
            messagebox.showinfo(MSG_SUCCESS, f"Backup realizado correctamente:\n{backup_file}")
            logging.info(f"Backup realizado: {backup_file}")
        except Exception as e:
            messagebox.showerror(MSG_ERROR, f"Error al realizar backup: {e}")
            logging.error(f"Error en backup: {e}")

    def restore_database(self):
        # Restaura la base de datos desde un archivo
        backup_file = filedialog.askopenfilename(
            title="Seleccionar archivo de backup",
            filetypes=[("Database files", "*.db"), ("All Files", "*.*")]
        )
        if backup_file:
            try:
                restore_database(backup_file)
                messagebox.showinfo(MSG_SUCCESS, "Base de datos restaurada correctamente.")
                logging.info(f"Base de datos restaurada desde: {backup_file}")
            except Exception as e:
                messagebox.showerror(MSG_ERROR, f"Error al restaurar la base de datos: {e}")
                logging.error(f"Error al restaurar DB: {e}")

    def run(self):
        # Inicia la aplicación y refresca la lista inicial
        self.refrescar_lista()
        self.root.mainloop()